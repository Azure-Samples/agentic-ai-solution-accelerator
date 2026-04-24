"""Generate docs/discovery/roi-calculator.xlsx.

One-shot builder for the partner-facing ROI worksheet shipped with the
accelerator. Run once at commit time, then re-run whenever the template
layout changes. The produced .xlsx is partner-editable; blue inputs drive
black formulas across the workbook.

Usage: python scripts/build-roi-calculator.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BLUE = Font(color="0000FF", name="Arial", size=11)
BLACK = Font(color="000000", name="Arial", size=11)
GREEN = Font(color="008000", name="Arial", size=11)
HEADER = Font(color="FFFFFF", name="Arial", size=11, bold=True)
BOLD = Font(color="000000", name="Arial", size=11, bold=True)
TITLE = Font(color="000000", name="Arial", size=14, bold=True)
YELLOW = PatternFill("solid", start_color="FFFF00")
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
SUBHEAD_FILL = PatternFill("solid", start_color="D9E1F2")
CURRENCY = '$#,##0;($#,##0);"-"'
CURRENCY2 = '$#,##0.00;($#,##0.00);"-"'
INT_FMT = '#,##0;(#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
MONTHS_FMT = '#,##0.0 "mo"'
YEARS_FMT = '#,##0.0 "yr"'
THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(cell):
    cell.font = HEADER
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_subhead(cell):
    cell.font = BOLD
    cell.fill = SUBHEAD_FILL


def build_readme(sheet):
    sheet.title = "README"
    sheet.column_dimensions["A"].width = 110
    rows = [
        ("Discovery ROI calculator", TITLE),
        ("", None),
        ("Partner-editable workbook that turns discovery answers into a signed-off ROI "
         "hypothesis and a copy-guide for the kpis[] block in accelerator.yaml.", None),
        ("", None),
        ("How to use:", BOLD),
        ("  1. Open 'Inputs' and fill every BLUE cell. YELLOW-highlighted cells are must-fill.", None),
        ("  2. Read 'ROI' for annual savings, payback, and 3-year cumulative impact.", None),
        ("  3. Read 'KPIs' — rows 5-7 render baseline/target values; hand-copy them into", None),
        ("     accelerator.yaml:kpis[] using the template in row 14. (Not paste-ready —", None),
        ("     Excel does not expand cell addresses into text when you copy out of a cell.)", None),
        ("  4. Attach this file to docs/discovery/solution-brief.md §4 as the ROI hypothesis source.", None),
        ("", None),
        ("Conventions (industry-standard financial-model colors):", BOLD),
        ("  BLUE cells       = hardcoded inputs you edit per customer", None),
        ("  BLACK cells      = formulas (do not hand-edit)", None),
        ("  GREEN cells      = links from other sheets in this workbook", None),
        ("  YELLOW fill      = must-fill assumption; workbook returns 0 until filled", None),
        ("", None),
        ("Honesty notes:", BOLD),
        ("  * This is a first-pass ROI *hypothesis*, not a signed business case. Expect to", None),
        ("    update baseline/target numbers after pilot evals run against golden cases.", None),
        ("  * Cost-avoidance and throughput gains are modeled; revenue lift is modeled but", None),
        ("    typically needs customer attribution data the partner does not have at discovery.", None),
        ("  * Azure run-cost is a placeholder input on the 'Inputs' sheet. Partners refine it", None),
        ("    using the Azure pricing calculator once architecture is confirmed during scaffold.", None),
        ("  * The KPIs sheet renders rows that mirror the accelerator.yaml:kpis schema", None),
        ("    (name / type / baseline / target). The 'type' column accepts duration_ms or ratio", None),
        ("    today because those are the only types wired into src/accelerator_baseline/telemetry.py.", None),
        ("", None),
        ("Related artifacts:", BOLD),
        ("  * docs/discovery/use-case-canvas.md       — 1-page exec alignment, before workshop", None),
        ("  * docs/discovery/discovery-workbook.csv   — structured workshop capture", None),
        ("  * docs/discovery/solution-brief.md        — canonical output of /discover-scenario", None),
        ("  * docs/discovery/SOLUTION-BRIEF-GUIDE.md  — how to run the workshop", None),
        ("  * docs/discovery/how-to-use.md            — sequencing across all five artifacts", None),
    ]
    for i, (text, font) in enumerate(rows, start=1):
        cell = sheet.cell(row=i, column=1, value=text)
        if font is not None:
            cell.font = font
        else:
            cell.font = BLACK
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_inputs(sheet):
    sheet.title = "Inputs"
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 60

    sheet["A1"] = "Inputs"
    sheet["A1"].font = TITLE
    sheet["A2"] = ("Edit BLUE cells. Units are noted per row. Cells with YELLOW fill "
                   "are must-fill; leaving them blank leaves ROI at zero.")
    sheet["A2"].font = BLACK
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    header_row = 4
    for col_idx, text in enumerate(["Input", "Value", "Unit", "Notes"], start=1):
        c = sheet.cell(row=header_row, column=col_idx, value=text)
        _style_header(c)

    # rows: (label, value, unit, note, must_fill)
    rows = [
        ("1. Engagement framing", None, None, None, None),
        ("Customer name", "<Customer>", "text",
         "Used on every sheet header. Partner fills during kickoff.", True),
        ("Engagement start date", None, "date",
         "Anchor for 1Y / 3Y savings projections.", True),
        ("Analysis horizon", 3, "years",
         "Default 3 years; edit if customer's business case runs longer.", False),
        ("Discount rate (annual)", 0.08, "ratio",
         "WACC or partner-standard hurdle rate. 0.08 = 8%.", False),

        ("2. Manual status-quo cost", None, None, None, None),
        ("FTEs doing this work today", None, "count",
         "Integer count. Include only FTEs whose time the agent displaces.", True),
        ("Fully-loaded annual cost per FTE", None, "$/yr",
         "Base + benefits + overhead. Use customer HR figure if shared; else market proxy.", True),
        ("% of FTE time spent on this process", 0.5, "ratio",
         "0.0 – 1.0. Default 0.5 = half of each person's time is on this process.", False),
        ("Transactions per year (if different denominator)", 0, "count",
         "Leave 0 if FTE-based. Only filled if cost model is $/transaction × volume.", False),
        ("Cost per transaction (manual, if using that model)", 0, "$/txn",
         "Leave 0 if FTE-based. Pairs with the row above.", False),

        ("3. Agent target performance", None, None, None, None),
        ("Time per task today (manual)", None, "minutes",
         "Median, not best case. Feeds briefing_production_time KPI.", True),
        ("Time per task with agent (target)", None, "minutes",
         "Target after GA. Must be achievable under acceptance.p95_latency_ms.", True),
        ("Target % of cases agent handles end-to-end", 0.8, "ratio",
         "The rest stay human or route through HITL. 0.0 – 1.0.", False),
        ("HITL approval rate (target)", 0.9, "ratio",
         "Fraction of agent drafts expected to ship as-is after review. Mirrors accelerator.yaml:kpis.hitl_approval_rate.", False),

        ("4. Azure + agent run cost", None, None, None, None),
        ("Estimated Azure run cost (monthly)", None, "$/mo",
         "Placeholder. Refine using Azure pricing calculator after /configure-landing-zone.", True),
        ("Cost per agent call (target)", 0.40, "$/call",
         "Must be ≤ accelerator.yaml:acceptance.cost_per_call_usd. Default matches shipped 0.40.", False),
        ("Monthly agent call volume", None, "calls/mo",
         "Expected steady-state call volume post-GA. Drives run-cost sanity check.", True),

        ("5. Optional revenue lift", None, None, None, None),
        ("Target revenue lift (annual, if attributable)", 0, "$/yr",
         "Often unknown at discovery. Leave 0 unless customer has attribution data.", False),
    ]

    r = header_row + 1
    for label, value, unit, note, must in rows:
        if value is None and unit is None and note is None:
            # section header
            cell = sheet.cell(row=r, column=1, value=label)
            _style_subhead(cell)
            for c in range(2, 5):
                sheet.cell(row=r, column=c).fill = SUBHEAD_FILL
            r += 1
            continue
        sheet.cell(row=r, column=1, value=label).font = BLACK
        v_cell = sheet.cell(row=r, column=2, value=value)
        v_cell.font = BLUE
        if unit in ("ratio",):
            v_cell.number_format = PCT_FMT
        elif unit in ("$/yr", "$/txn"):
            v_cell.number_format = CURRENCY
        elif unit in ("$/call", "$/mo"):
            v_cell.number_format = CURRENCY2
        elif unit in ("minutes", "count", "calls/mo", "years"):
            v_cell.number_format = INT_FMT if unit != "minutes" else '#,##0.0;(#,##0.0);"-"'
        if must and (value is None or value == 0 and unit != "ratio"):
            v_cell.fill = YELLOW
        sheet.cell(row=r, column=3, value=unit).font = BLACK
        n_cell = sheet.cell(row=r, column=4, value=note)
        n_cell.font = BLACK
        n_cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # Define a name map for formula readability via a small dictionary on ROI sheet
    return {
        "customer_name": "Inputs!$B$6",
        "start_date": "Inputs!$B$7",
        "horizon_yrs": "Inputs!$B$8",
        "discount_rate": "Inputs!$B$9",
        "fte_count": "Inputs!$B$11",
        "fte_cost": "Inputs!$B$12",
        "fte_pct": "Inputs!$B$13",
        "txn_vol": "Inputs!$B$14",
        "txn_cost": "Inputs!$B$15",
        "time_manual": "Inputs!$B$17",
        "time_agent": "Inputs!$B$18",
        "coverage": "Inputs!$B$19",
        "hitl_rate": "Inputs!$B$20",
        "azure_monthly": "Inputs!$B$22",
        "cost_per_call": "Inputs!$B$23",
        "calls_monthly": "Inputs!$B$24",
        "rev_lift": "Inputs!$B$26",
    }


def build_roi(sheet, refs):
    sheet.title = "ROI"
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 48

    sheet["A1"] = "ROI model"
    sheet["A1"].font = TITLE
    sheet["A2"] = (
        'Formulas reference "Inputs" (GREEN = cross-sheet link). Revisit after pilot to '
        "reconcile against measured performance.")
    sheet["A2"].font = BLACK

    # Section A: status-quo annual cost
    sheet["A4"] = "A. Manual status-quo annual cost"
    _style_subhead(sheet["A4"])
    sheet["A5"] = "FTE-based component"
    sheet["B5"] = f"={refs['fte_count']}*{refs['fte_cost']}*{refs['fte_pct']}"
    sheet["A6"] = "Transaction-based component"
    sheet["B6"] = f"={refs['txn_vol']}*{refs['txn_cost']}"
    sheet["A7"] = "Total manual cost (annual)"
    sheet["B7"] = "=B5+B6"
    for r in (5, 6, 7):
        sheet.cell(row=r, column=2).number_format = CURRENCY
        sheet.cell(row=r, column=2).font = GREEN if r < 7 else BOLD
    sheet["B7"].font = BOLD

    # Section B: agent annual run cost
    sheet["A9"] = "B. Agent annual run cost"
    _style_subhead(sheet["A9"])
    sheet["A10"] = "Annual Azure run cost (infra)"
    sheet["B10"] = f"={refs['azure_monthly']}*12"
    sheet["A11"] = "Annual agent call cost (LLM + tool)"
    sheet["B11"] = f"={refs['calls_monthly']}*12*{refs['cost_per_call']}"
    sheet["A12"] = "Total agent run cost (annual)"
    sheet["B12"] = "=B10+B11"
    for r in (10, 11, 12):
        sheet.cell(row=r, column=2).number_format = CURRENCY
        sheet.cell(row=r, column=2).font = GREEN if r < 12 else BOLD
    sheet["B12"].font = BOLD

    # Section C: annual savings + net impact
    sheet["A14"] = "C. Annual impact"
    _style_subhead(sheet["A14"])
    sheet["A15"] = "Labor savings (coverage × manual cost)"
    sheet["B15"] = f"={refs['coverage']}*B7"
    sheet["A16"] = "Revenue lift (input)"
    sheet["B16"] = f"={refs['rev_lift']}"
    sheet["A17"] = "Gross annual benefit"
    sheet["B17"] = "=B15+B16"
    sheet["A18"] = "Less: agent run cost"
    sheet["B18"] = "=-B12"
    sheet["A19"] = "Net annual impact"
    sheet["B19"] = "=B17+B18"
    for r in (15, 16, 17, 18, 19):
        sheet.cell(row=r, column=2).number_format = CURRENCY
        sheet.cell(row=r, column=2).font = BLACK if r < 19 else BOLD
    sheet["B19"].font = BOLD

    # Section D: payback + NPV
    sheet["A21"] = "D. Payback & NPV"
    _style_subhead(sheet["A21"])
    sheet["A22"] = "One-time implementation cost (optional)"
    sheet["B22"] = 0
    sheet["B22"].font = BLUE
    sheet["B22"].number_format = CURRENCY
    sheet["C22"] = "edit if partner has a fixed-fee component"
    sheet["C22"].font = BLACK

    sheet["A23"] = "Simple payback (months)"
    # Distinguish three states:
    #   B19 blank/zero (inputs not filled) -> "fill inputs"
    #   B19 > 0 (positive impact)           -> payback months
    #   B19 < 0 (negative impact)           -> "no payback at current inputs"
    sheet["B23"] = '=IFERROR(IF(ISBLANK(B19),"fill inputs",IF(B19=0,"fill inputs",IF(B19<0,"no payback at current inputs",B22/B19*12))),"fill inputs")'
    sheet["B23"].number_format = MONTHS_FMT
    sheet["C23"] = "Shows 'no payback at current inputs' if annual impact is negative."
    sheet["C23"].font = BLACK

    sheet["A24"] = "3-year cumulative net impact"
    sheet["B24"] = f"=B19*MIN({refs['horizon_yrs']},3)-B22"
    sheet["B24"].number_format = CURRENCY
    sheet["B24"].font = BOLD

    sheet["A25"] = "NPV over horizon"
    sheet["B25"] = (
        f"=IFERROR(NPV({refs['discount_rate']},"
        f"INDEX(ROW(INDIRECT(\"1:\"&MAX({refs['horizon_yrs']},1))),0,0)*0+B19)-B22,NA())"
    )
    # fallback simpler NPV: sum discounted constant cashflows over horizon
    sheet["B25"] = (
        f"=IFERROR(B19*((1-(1+{refs['discount_rate']})^-{refs['horizon_yrs']})/{refs['discount_rate']})-B22,\"fill inputs\")"
    )
    sheet["B25"].number_format = CURRENCY

    # Section E: yearly schedule (horizon-aware up to 5)
    sheet["A27"] = "E. Yearly cashflow schedule (up to 5 years)"
    _style_subhead(sheet["A27"])
    headers = ["Year", "Net impact", "Discount factor", "PV of year", "Cumulative PV"]
    for i, h in enumerate(headers, start=1):
        c = sheet.cell(row=28, column=i, value=h)
        _style_header(c)
    for y in range(1, 6):
        r = 28 + y
        sheet.cell(row=r, column=1, value=f"Year {y}").font = BLACK
        # include year only if within horizon
        sheet.cell(row=r, column=2,
                   value=f"=IF({refs['horizon_yrs']}>={y},B19,0)").number_format = CURRENCY
        sheet.cell(row=r, column=3,
                   value=f"=IF({refs['horizon_yrs']}>={y},1/(1+{refs['discount_rate']})^{y},0)"
                   ).number_format = '0.000'
        sheet.cell(row=r, column=4,
                   value=f"=B{r}*C{r}").number_format = CURRENCY
        if y == 1:
            sheet.cell(row=r, column=5, value=f"=D{r}-B22").number_format = CURRENCY
        else:
            sheet.cell(row=r, column=5, value=f"=E{r-1}+D{r}").number_format = CURRENCY


def build_kpis(sheet, refs):
    sheet.title = "KPIs"
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 48

    sheet["A1"] = "KPI copy-guide — hand-copy values from rows 5-7 into accelerator.yaml:kpis"
    sheet["A1"].font = TITLE
    sheet["A2"] = (
        "Mirrors the kpis[] schema: {name, type, baseline, target}. Only duration_ms and "
        "ratio are wired into src/accelerator_baseline/telemetry.py today; other types "
        "require partner code.")
    sheet["A2"].font = BLACK
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    for i, h in enumerate(["KPI event name", "type", "baseline", "target", "Notes"], start=1):
        c = sheet.cell(row=4, column=i, value=h)
        _style_header(c)

    kpis = [
        ("briefing_production_time", "duration_ms",
         f"={refs['time_manual']}*60*1000", f"={refs['time_agent']}*60*1000",
         "Time from request.received to response.returned. Sourced from Inputs §3."),
        ("hitl_approval_rate", "ratio",
         1.0, f"={refs['hitl_rate']}",
         "Fraction of drafts approved as-is. Baseline is 1.0 (fully manual)."),
        ("coverage_rate", "ratio",
         0.0, f"={refs['coverage']}",
         "Fraction of volume agent handles end-to-end. Remainder routes to HITL or human."),
    ]
    for i, (name, typ, base, tgt, note) in enumerate(kpis, start=5):
        sheet.cell(row=i, column=1, value=name).font = BLACK
        sheet.cell(row=i, column=2, value=typ).font = BLACK
        b_cell = sheet.cell(row=i, column=3, value=base)
        t_cell = sheet.cell(row=i, column=4, value=tgt)
        if typ == "duration_ms":
            b_cell.number_format = INT_FMT
            t_cell.number_format = INT_FMT
            b_cell.font = GREEN
            t_cell.font = GREEN
        elif typ == "ratio":
            b_cell.number_format = PCT_FMT
            t_cell.number_format = PCT_FMT
            b_cell.font = GREEN if isinstance(base, str) and base.startswith("=") else BLUE
            t_cell.font = GREEN
        n_cell = sheet.cell(row=i, column=5, value=note)
        n_cell.font = BLACK
        n_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Copy-guide (NOT paste-ready) — partner hand-copies values from row 5-7 above.
    sheet["A11"] = "Copy guide — hand-copy the numeric values from rows 5-7 above"
    _style_subhead(sheet["A11"])
    sheet["A12"] = (
        "After you fill Inputs, each row above renders a baseline and target value. "
        "Hand-copy those numbers into accelerator.yaml:kpis using the template below. "
        "This is NOT a paste-ready block — Excel does not expand cell addresses into "
        "values when you copy text out of a cell."
    )
    sheet["A12"].font = BLACK
    sheet["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[12].height = 45

    yaml_template = (
        "kpis:\n"
        "  - name: briefing_production_time\n"
        "    type: duration_ms\n"
        "    baseline: <copy C5 as integer ms>\n"
        "    target:   <copy D5 as integer ms>\n"
        "  - name: hitl_approval_rate\n"
        "    type: ratio\n"
        "    baseline: 1.0\n"
        "    target:   <copy D6 as decimal, e.g. 0.8>\n"
        "  - name: coverage_rate\n"
        "    type: ratio\n"
        "    baseline: 0.0\n"
        "    target:   <copy D7 as decimal, e.g. 0.7>\n"
        "\n"
        "# cost_per_call_usd belongs in accelerator.yaml:acceptance, not kpis[].\n"
        "# acceptance.cost_per_call_usd: <copy Inputs!B23 as decimal>\n"
    )
    sheet["A14"] = yaml_template
    sheet["A14"].font = Font(name="Consolas", size=10)
    sheet["A14"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[14].height = 260


def main():
    wb = Workbook()
    readme = wb.active
    build_readme(readme)
    inputs = wb.create_sheet("Inputs")
    refs = build_inputs(inputs)
    roi = wb.create_sheet("ROI")
    build_roi(roi, refs)
    kpis = wb.create_sheet("KPIs")
    build_kpis(kpis, refs)

    out = Path(__file__).resolve().parent.parent / "docs" / "discovery" / "roi-calculator.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
